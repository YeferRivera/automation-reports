import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

class ReportGenerator:
    def __init__(self):
        self.beige_fill = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")
        self.light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        self.bold_font = Font(bold=True)
    
    def generate_summary_report(self, resumen_data, output_path="resume.xlsx"):
        """Genera el reporte de resumen con formato"""
        # Guardar el DataFrame a Excel
        resumen_data.to_excel(output_path, index=False)
        
        # Aplicar formato
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Formato para encabezados
        for cell in ws[1]:
            cell.fill = self.beige_fill
            cell.font = self.bold_font
        
        # Formato para filas de total
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[2].value and isinstance(row[2].value, str) and row[2].value.startswith("Total -"):
                for cell in row:
                    cell.fill = self.light_blue_fill
        
        # Formato de porcentaje
        promedio_col_idx = None
        for idx, cell in enumerate(ws[1]):
            if cell.value == 'Promedio':
                promedio_col_idx = idx
                break
        
        if promedio_col_idx is not None:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=promedio_col_idx + 1)
                if cell.value is not None:
                    cell.number_format = '0"%"'
        
        # Crear tabla simplificada
        self._crear_tabla_simplificada(wb, resumen_data)
        
        # Guardar con formato
        wb.save(output_path)
        return output_path
    
    def _crear_tabla_simplificada(self, wb, resumen_data):
        """Crea una hoja adicional con tabla simplificada"""
        # Crear nueva hoja
        if "Tabla Simplificada" in wb.sheetnames:
            ws_tabla = wb["Tabla Simplificada"]
            wb.remove(ws_tabla)
        ws_tabla = wb.create_sheet("Tabla Simplificada")
        
        # Filtrar filas de total
        datos_filtrados = resumen_data[~resumen_data['Número de Serie de la Impresora'].astype(str).str.startswith('Total -')]
        
        # Eliminar columnas no necesarias
        columnas_a_eliminar = ['Páginas', 'Factura', 'Promedio']
        datos_sin_columnas = datos_filtrados.drop(columns=columnas_a_eliminar)
        
        # Renombrar Ciudad a Municipio
        datos_sin_columnas = datos_sin_columnas.rename(columns={'Ciudad': 'Municipio'})
        
        # Reorganizar columnas
        columnas_geo = ['Municipio', 'Ubicación', 'Unidad de Negocio']
        otras_columnas = [col for col in datos_sin_columnas.columns if col not in columnas_geo]
        orden_final = columnas_geo + otras_columnas
        
        # Crear dataframe final
        datos_simplificados = datos_sin_columnas[orden_final]
        
        # Escribir encabezados
        for col_idx, column_name in enumerate(datos_simplificados.columns, 1):
            cell = ws_tabla.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.fill = self.beige_fill
            cell.font = self.bold_font
        
        # Escribir datos
        for row_idx, row_data in enumerate(datos_simplificados.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_tabla.cell(row=row_idx, column=col_idx).value = value
        
        # Determinar rango de tabla
        num_filas = len(datos_simplificados) + 1
        num_columnas = len(datos_simplificados.columns)
        rango_tabla = f"A1:{chr(64 + num_columnas)}{num_filas}"
        
        # Crear tabla con estilo
        tabla = Table(displayName="TablaSimplificada", ref=rango_tabla)
        estilo = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo
        
        # Agregar tabla a la hoja
        ws_tabla.add_table(tabla)
        
        # Auto-ajustar columnas
        for col in ws_tabla.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            
            adjusted_width = (max_length + 2) * 1.2
            ws_tabla.column_dimensions[column].width = adjusted_width