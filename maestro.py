import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import numpy as np
from openpyxl.utils import get_column_letter, column_index_from_string

# Cargar el archivo maestro
maestro_df = pd.read_csv("MAESTRO_ENERO.csv", sep=";", encoding="ISO-8859-1", low_memory=False)

# === Cargar archivo de factura y extraer nombre de unidad de negocio ===
# Primero extraemos el nombre de la unidad de negocio usando openpyxl
wb_factura = openpyxl.load_workbook("factura_enero.xlsx", data_only=True)
ws_factura = wb_factura.active

# Diccionario para almacenar la relación entre números de serie y unidades de negocio
serie_unidad_negocio = {}
serie_ciudad = {}       # Nuevo diccionario para Ciudad
serie_ubicacion = {}    # Nuevo diccionario para Ubicación

# Lista para almacenar todas las unidades de negocio encontradas
unidades_negocio = []

# Primero identificar todas las unidades de negocio en celdas combinadas
for merged_range in ws_factura.merged_cells.ranges:
    min_col, min_row, max_col, max_row = merged_range.bounds
    # Celdas combinadas que abarcan columnas C y D
    if min_col == 3 and max_col == 4:
        cell_value = ws_factura.cell(row=min_row, column=min_col).value
        if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 0:
            unidades_negocio.append({
                'nombre': cell_value.strip(),
                'fila': min_row
            })
            print(f"Unidad de Negocio encontrada en fila {min_row}: {cell_value.strip()}")

# Ordenar unidades de negocio por número de fila para procesar en orden
unidades_negocio.sort(key=lambda x: x['fila'])

# Ahora cargar los datos normales de factura para asignar números de serie a unidades, ciudad y ubicación
factura_df = pd.read_excel("factura_enero.xlsx", header=None)

# Crear un DataFrame con columnas: Número de Serie, Factura, Ciudad, Ubicación y fila original
factura_data_con_fila = pd.DataFrame()
factura_data_con_fila['Número de Serie'] = factura_df.iloc[:, 4]  # Columna E (índice 4)
factura_data_con_fila['Factura'] = factura_df.iloc[:, 12]  # Columna M (índice 12)
factura_data_con_fila['Ciudad'] = factura_df.iloc[:, 6]    # Columna G (índice 6)
factura_data_con_fila['Ubicación'] = factura_df.iloc[:, 8]  # Columna I (índice 8)
factura_data_con_fila['fila'] = factura_data_con_fila.index + 1  # Índice base 1 para coincidir con Excel

# Eliminar filas sin número de serie
factura_data_con_fila = factura_data_con_fila.dropna(subset=['Número de Serie'])

# Asignar unidad de negocio, ciudad y ubicación a cada número de serie
for i, serie_row in factura_data_con_fila.iterrows():
    serie = serie_row['Número de Serie']
    fila = serie_row['fila']
    ciudad = serie_row['Ciudad']
    ubicacion = serie_row['Ubicación']
    
    # Guardar ciudad y ubicación (si no son nulos)
    if pd.notna(ciudad):
        serie_ciudad[str(serie)] = ciudad
    if pd.notna(ubicacion):
        serie_ubicacion[str(serie)] = ubicacion
    
    # Encontrar la unidad de negocio correspondiente (la última cuya fila es menor a la del número de serie)
    unidad_asignada = None
    for unidad in reversed(unidades_negocio):
        if unidad['fila'] < fila:
            unidad_asignada = unidad['nombre']
            break
    
    if unidad_asignada:
        serie_unidad_negocio[str(serie)] = unidad_asignada

# Preparar datos de factura para el merge
factura_data = factura_data_con_fila[['Número de Serie', 'Factura']].copy()
factura_data.columns = ['Número de Serie de la Impresora', 'Factura']

# Columnas por las que agruparemos
columnas_grupo = [
    'Nombre de la impresora',
    'Tipo/Modelo impresora',
    'Número de Serie de la Impresora',
    'Nombre Completo',
    'Nombre de Usuario',
    'Departamento del Usuario'
]

# Agrupar y sumar
resumen_df = maestro_df.groupby(columnas_grupo, as_index=False)['Páginas'].sum()

# === Cruce por Número de Serie de la Impresora ===
resumen_con_factura = pd.merge(
    resumen_df,
    factura_data,
    on='Número de Serie de la Impresora',
    how='left'  # Mantener todo lo del resumen aunque falte en factura
)

# Calcular totales por impresora para usarlos en los porcentajes
totales_impresora = {}
facturas_impresora = {}
for impresora, grupo in resumen_con_factura.groupby('Nombre de la impresora'):
    totales_impresora[impresora] = grupo['Páginas'].sum()
    # Guardar el primer valor no nulo de factura para esta impresora
    factura_valores = grupo['Factura'].dropna()
    if not factura_valores.empty:
        facturas_impresora[impresora] = factura_valores.iloc[0]
    else:
        facturas_impresora[impresora] = None

# Crear lista con totales por impresora incluyendo factura
final_data = []
for impresora, grupo in resumen_con_factura.groupby('Nombre de la impresora'):
    # Obtener el total de páginas y valor de factura para esta impresora
    total_paginas_impresora = totales_impresora[impresora]
    factura_valor = facturas_impresora.get(impresora)
    
    # Obtener la unidad de negocio, ciudad y ubicación para este grupo
    numero_serie = str(grupo['Número de Serie de la Impresora'].iloc[0])
    unidad_negocio = serie_unidad_negocio.get(numero_serie, "No identificada")
    ciudad = serie_ciudad.get(numero_serie, "No identificada")
    ubicacion = serie_ubicacion.get(numero_serie, "No identificada")
    
    # Añadir filas de detalle con promedio y total páginas
    grupo_sin_factura = grupo.copy()
    grupo_sin_factura['Factura'] = None  # Eliminar factura de filas de detalle
    
    # Calcular promedio exacto (con todos los decimales) para cálculos internos
    grupo_sin_factura['Promedio_exacto'] = grupo_sin_factura['Páginas'] / total_paginas_impresora * 100
    
    # Calcular promedio redondeado para visualización
    grupo_sin_factura['Promedio'] = grupo_sin_factura['Promedio_exacto'].round().astype(int)
    
    # Calcular total páginas usando el promedio exacto (con todos los decimales)
    if factura_valor is not None:
        grupo_sin_factura['Total paginas'] = ((grupo_sin_factura['Promedio_exacto'] / 100) * factura_valor).round().astype(int)
    else:
        grupo_sin_factura['Total paginas'] = None
    
    # Eliminar la columna de promedio exacto antes de añadir al resultado final
    grupo_sin_factura = grupo_sin_factura.drop(columns=['Promedio_exacto'])
    
    # Añadir columnas de información adicional
    grupo_sin_factura['Unidad de Negocio'] = unidad_negocio
    grupo_sin_factura['Ciudad'] = ciudad
    grupo_sin_factura['Ubicación'] = ubicacion
    
    final_data.append(grupo_sin_factura)
    
    # Crear fila de total con factura
    total_row = {
        'Nombre de la impresora': impresora,
        'Tipo/Modelo impresora': '',
        'Número de Serie de la Impresora': f'Total - {grupo["Número de Serie de la Impresora"].iloc[0]}',
        'Nombre Completo': '',
        'Nombre de Usuario': '',
        'Departamento del Usuario': '',
        'Páginas': total_paginas_impresora,
        'Factura': factura_valor,
        'Promedio': 100,  # El total representa el 100%
        'Total paginas': factura_valor if factura_valor is not None else None,
        'Unidad de Negocio': unidad_negocio,
        'Ciudad': ciudad,
        'Ubicación': ubicacion
    }
    final_data.append(pd.DataFrame([total_row]))

# Concatenar resultados
resumen_con_totales = pd.concat(final_data, ignore_index=True)

# Comprobar si el archivo existe y eliminarlo si es necesario
output_file = "resumen_dinamico_formateado.xlsx"
if os.path.exists(output_file):
    try:
        os.remove(output_file)
    except PermissionError:
        output_file = "resumen_dinamico_formateado_new.xlsx"
        print(f"El archivo original está en uso. Guardando como: {output_file}")

# Guardar a Excel
resumen_con_totales.to_excel(output_file, index=False)

# Aplicar formato con openpyxl
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# Formato para títulos: beige + negrita
beige_fill = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.fill = beige_fill
    cell.font = bold_font

# Formato para filas de total: fondo AZUL CLARO
light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[2].value and isinstance(row[2].value, str) and row[2].value.startswith("Total -"):
        for cell in row:
            cell.fill = light_blue_fill

# Encontrar la columna de Promedio
header_row = list(ws.rows)[0]
promedio_col_idx = None
for idx, cell in enumerate(header_row):
    if cell.value == 'Promedio':
        promedio_col_idx = idx
        break

# Aplicar formato de porcentaje a la columna Promedio
if promedio_col_idx is not None:
    for row_idx in range(2, ws.max_row + 1):  # Empezar desde la fila 2 (después del encabezado)
        cell = ws.cell(row=row_idx, column=promedio_col_idx + 1)  # +1 porque openpyxl usa índices 1-based
        if cell.value is not None:
            cell.number_format = '0"%"'  # Formato de porcentaje sin decimales

# Guardar cambios de formato
wb.save(output_file)

# === Guardar archivo final ===
output_path = "resume.xlsx"
# Comprobar si el archivo existe y eliminarlo si es necesario
if os.path.exists(output_path):
    try:
        os.remove(output_path)
    except PermissionError:
        output_path = "resume_new.xlsx"
        print(f"El archivo resume.xlsx está en uso. Guardando como: {output_path}")

# Guardar a Excel y aplicar formato
resumen_con_totales.to_excel(output_path, index=False)

# Aplicar formato al archivo final
wb = openpyxl.load_workbook(output_path)
ws = wb.active

# Formato para títulos: beige + negrita
for cell in ws[1]:
    cell.fill = beige_fill
    cell.font = bold_font

# Formato para filas de total: fondo AZUL CLARO
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[2].value and isinstance(row[2].value, str) and row[2].value.startswith("Total -"):
        for cell in row:
            cell.fill = light_blue_fill

# Aplicar formato de porcentaje a la columna Promedio
if promedio_col_idx is not None:
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=promedio_col_idx + 1)
        if cell.value is not None:
            cell.number_format = '0"%"'  # Formato de porcentaje sin decimales

# Crear una nueva hoja con tabla simplificada
if "Tabla Simplificada" in wb.sheetnames:
    ws_tabla = wb["Tabla Simplificada"]
    wb.remove(ws_tabla)  # Eliminar si ya existe
ws_tabla = wb.create_sheet("Tabla Simplificada")

# Filtrar el dataframe para eliminar filas de total (aquellas que tienen "Total -" en el número de serie)
datos_filtrados = resumen_con_totales[~resumen_con_totales['Número de Serie de la Impresora'].astype(str).str.startswith('Total -')]

# Eliminar las columnas especificadas
columnas_a_eliminar = ['Páginas', 'Factura', 'Promedio']
datos_sin_columnas = datos_filtrados.drop(columns=columnas_a_eliminar)

# Renombrar Ciudad a Municipio
datos_sin_columnas = datos_sin_columnas.rename(columns={'Ciudad': 'Municipio'})

# Reorganizar las columnas para poner Municipio, Ubicación y Unidad de Negocio primero
columnas_geo = ['Municipio', 'Ubicación', 'Unidad de Negocio']
otras_columnas = [col for col in datos_sin_columnas.columns if col not in columnas_geo]
orden_final = columnas_geo + otras_columnas

# Crear el dataframe final con el orden de columnas deseado
datos_simplificados = datos_sin_columnas[orden_final]

# Escribir los encabezados en la nueva hoja
for col_idx, column_name in enumerate(datos_simplificados.columns, 1):
    cell = ws_tabla.cell(row=1, column=col_idx)
    cell.value = column_name
    cell.fill = beige_fill
    cell.font = bold_font

# Escribir los datos en la nueva hoja
for row_idx, row_data in enumerate(datos_simplificados.values, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_tabla.cell(row=row_idx, column=col_idx).value = value

# Determinar el rango de la tabla
num_filas = len(datos_simplificados) + 1  # +1 para incluir la fila de encabezado
num_columnas = len(datos_simplificados.columns)
rango_tabla = f"A1:{chr(64 + num_columnas)}{num_filas}"

# Crear una tabla con estilo
tabla = Table(displayName="TablaSimplificada", ref=rango_tabla)

# Definir el estilo de la tabla (usando un estilo predefinido de Excel)
estilo = TableStyleInfo(
    name="TableStyleMedium9", 
    showFirstColumn=False,
    showLastColumn=False, 
    showRowStripes=True, 
    showColumnStripes=False
)
tabla.tableStyleInfo = estilo

# Agregar la tabla a la hoja
ws_tabla.add_table(tabla)

# Auto-ajustar ancho de columnas
for col in ws_tabla.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            cell_len = len(str(cell.value))
            if cell_len > max_length:
                max_length = cell_len
    
    adjusted_width = (max_length + 2) * 1.2
    ws_tabla.column_dimensions[column].width = adjusted_width

# Guardar los cambios
wb.save(output_path)
print(f"Archivo guardado correctamente como {output_path} con la nueva hoja 'Tabla Simplificada'")
print(f"Se encontraron {len(unidades_negocio)} unidades de negocio")
for unidad in unidades_negocio:
    print(f"- {unidad['nombre']}")